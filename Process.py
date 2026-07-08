from pathlib import Path

from openpyxl import load_workbook
from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QStackedWidget,
    QVBoxLayout,
    QWidget,
)

from Method import Method


EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
ACTION_NAMES = ("文件读取", "文件复制", "文件移动", "文件替换", "文件改名", "文件删除")
# 统一放在这里，避免每个控件单独重复设置样式。
MODERN_STYLE = """
QPushButton {
    min-height: 38px;
    padding: 7px 18px;
    border-radius: 12px;
    border: 1px solid #cfd7e3;
    background-color: #f7f9fc;
    color: #1f2937;
    font-size: 14px;
    font-weight: 600;
}
QPushButton:hover {
    background-color: #eef3f9;
    border-color: #b8c4d4;
}
QPushButton:pressed {
    background-color: #e2e9f2;
}
QPushButton#yellowButton {
    background-color: #f7c948;
    border-color: #e1ad19;
    color: #2f2600;
}
QPushButton#yellowButton:hover {
    background-color: #f5bd24;
}
QPushButton#yellowButton:pressed {
    background-color: #e0a80f;
}
QPushButton#greenButton {
    background-color: #21a67a;
    border-color: #15805e;
    color: #ffffff;
}
QPushButton#greenButton:hover {
    background-color: #15956a;
}
QPushButton#greenButton:pressed {
    background-color: #0f7a55;
}
QLineEdit, QComboBox {
    min-height: 34px;
    padding: 5px 10px;
    border-radius: 8px;
    border: 1px solid #cfd7e3;
    background-color: #ffffff;
}
"""


class ProcessFrame(QWidget):
    """主界面：负责配置展示、配置编辑和把按钮操作分发到 Method。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(MODERN_STYLE)

        self.config = Method.readData()
        self._sync_config_fields()

        # 这些控件会在页面构建时创建，先占位便于其它方法安全访问。
        self.start_path_edit = None
        self.target_path_edit = None
        self.excel_edit = None
        self.sheet_combo = None
        self.display_value_labels = {}
        self.file_column_edit = None
        self.original_column_edit = None
        self.new_column_edit = None
        self.output_column_edit = None

        self.include_subdirs_check = QCheckBox("包含子目录")
        self.folder_only_check = QCheckBox("仅修改文件夹")
        self.action_combo = QComboBox()
        self.action_combo.addItems(ACTION_NAMES)
        self.action_combo.setCurrentText("文件读取")
        self.action_combo.setFixedWidth(120)

        self.stack = QStackedWidget()
        self.show_page = QWidget()
        self.input_page = QWidget()
        # 两个页面共用同一个窗口：展示页和修改配置页。
        self.stack.addWidget(self.show_page)
        self.stack.addWidget(self.input_page)

        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(24, 18, 24, 18)
        root_layout.addWidget(self.stack)

        self._build_show_page()
        self._build_input_page()
        self.showConfigPage()

    def _sync_config_fields(self):
        # 把持久化配置同步成界面显示和业务方法直接使用的字段。
        self.start_path_data = self.config.get("start_path", "")
        self.target_path_data = self.config.get("target_path", "")
        self.excel_data = self.config.get("excel", "")
        self.sheet_data = self.config.get("sheetname", "Sheet1")
        self.file_column_data = self.config.get("file_column", self.config.get("start", "1"))
        self.original_column_data = self.config.get("original_column", "2")
        self.new_column_data = self.config.get("new_column", self.config.get("target", "3"))
        self.output_column_data = self.config.get("output_column", "4")

    def _desktop_path(self):
        desktop = Path.home() / "Desktop"
        if desktop.is_dir():
            return str(desktop)
        return str(Path.home())

    def _load_sheet_names(self, excel_path):
        workbook = load_workbook(excel_path, read_only=True)
        try:
            return workbook.sheetnames
        finally:
            workbook.close()

    def _add_display_row(self, layout, row, label_text, value_text):
        label = QLabel(label_text)
        label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        label.setMinimumWidth(110)

        value = QLabel(value_text or "")
        value.setTextInteractionFlags(Qt.TextSelectableByMouse)
        value.setWordWrap(True)

        layout.addWidget(label, row, 0)
        layout.addWidget(value, row, 1)
        return value

    def _make_readonly_picker_row(self, label_text, edit, button_text, button_handler):
        # 路径输入统一走选择窗口，减少手输路径带来的错误。
        row = QHBoxLayout()
        label = QLabel(label_text)
        label.setMinimumWidth(110)
        label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        edit.setReadOnly(True)
        choose_button = QPushButton(button_text)
        choose_button.clicked.connect(button_handler)

        row.addWidget(label)
        row.addWidget(edit, 1)
        row.addWidget(choose_button)
        return row

    def _build_show_page(self):
        layout = QVBoxLayout(self.show_page)
        layout.setSpacing(12)

        data_layout = QGridLayout()
        data_layout.setColumnStretch(1, 1)
        layout.addLayout(data_layout)

        # 展示页每一行都保存 getter，配置更新后可以统一刷新。
        self.display_rows = [
            ("文件起始地址:", lambda: self.start_path_data),
            ("文件目标地址:", lambda: self.target_path_data),
            ("表格文件地址:", lambda: self.excel_data),
            ("工作表名称:", lambda: self.sheet_data),
            ("文件名列:", lambda: self.file_column_data),
            ("原名称列:", lambda: self.original_column_data),
            ("新名称列:", lambda: self.new_column_data),
            ("写入列:", lambda: self.output_column_data),
        ]

        for row, (label, getter) in enumerate(self.display_rows):
            self.display_value_labels[label] = self._add_display_row(data_layout, row, label, getter())

        data_button_row = QHBoxLayout()
        data_button_row.addStretch()
        modify_button = QPushButton("修改数据")
        modify_button.setObjectName("yellowButton")
        modify_button.setFixedWidth(150)
        modify_button.clicked.connect(self.inputDataButton)
        data_button_row.addWidget(modify_button)

        init_button = QPushButton("初始化数据表格")
        init_button.setObjectName("yellowButton")
        init_button.setFixedWidth(150)
        init_button.clicked.connect(self.initializeSheetButton)
        data_button_row.addWidget(init_button)
        data_button_row.addStretch()
        layout.addLayout(data_button_row)

        layout.addStretch()

        option_row = QHBoxLayout()
        option_row.addStretch()
        option_row.addWidget(self.include_subdirs_check)
        option_row.addWidget(self.folder_only_check)
        option_row.addStretch()
        layout.addLayout(option_row)

        action_row = QHBoxLayout()
        action_row.addStretch()

        execute_box = QVBoxLayout()
        execute_box.addWidget(self.action_combo)
        # “执行”使用原名称列 -> 新名称列；“全名称操作”用于文件名列 -> 写入列。
        execute_button = QPushButton("执行")
        execute_button.setObjectName("greenButton")
        execute_button.setFixedWidth(110)
        execute_button.clicked.connect(self.executeAction)
        execute_box.addWidget(execute_button)

        full_name_button = QPushButton("全名称操作")
        full_name_button.setObjectName("greenButton")
        full_name_button.setFixedWidth(110)
        full_name_button.clicked.connect(self.executeFullNameAction)
        execute_box.addWidget(full_name_button)
        action_row.addLayout(execute_box)

        action_row.addStretch()
        layout.addLayout(action_row)

    def _build_input_page(self):
        layout = QVBoxLayout(self.input_page)
        layout.setSpacing(10)

        self.start_path_edit = QLineEdit()
        self.target_path_edit = QLineEdit()
        self.excel_edit = QLineEdit()
        self.sheet_combo = QComboBox()
        self.sheet_combo.setEditable(False)
        self.file_column_edit = QLineEdit()
        self.original_column_edit = QLineEdit()
        self.new_column_edit = QLineEdit()
        self.output_column_edit = QLineEdit()

        layout.addLayout(
            self._make_readonly_picker_row(
                "文件起始地址:",
                self.start_path_edit,
                "选择",
                self.chooseStartFolder,
            )
        )
        layout.addLayout(
            self._make_readonly_picker_row(
                "文件目标地址:",
                self.target_path_edit,
                "选择",
                self.chooseTargetFolder,
            )
        )
        layout.addLayout(
            self._make_readonly_picker_row(
                "表格文件地址:",
                self.excel_edit,
                "选择",
                self.chooseExcelFile,
            )
        )

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignRight)
        form.addRow("工作表名称:", self.sheet_combo)
        form.addRow("文件名列:", self.file_column_edit)
        form.addRow("原名称列:", self.original_column_edit)
        form.addRow("新名称列:", self.new_column_edit)
        form.addRow("写入列:", self.output_column_edit)
        layout.addLayout(form)

        button_row = QHBoxLayout()
        button_row.addStretch()
        define_button = QPushButton("确认修改")
        define_button.clicked.connect(self.defineDataButton)
        back_button = QPushButton("返回")
        back_button.clicked.connect(self.backButton)
        button_row.addWidget(define_button)
        button_row.addWidget(back_button)
        button_row.addStretch()
        layout.addLayout(button_row)
        layout.addStretch()

    def _refresh_show_page(self):
        # 从当前配置重新填充展示页，避免修改后仍显示旧内容。
        for label, getter in self.display_rows:
            value_label = self.display_value_labels.get(label)
            if value_label is not None:
                value_label.setText(getter() or "")

    def _file_names_from_excel(self):
        # 复制、移动、替换、删除都以文件名列作为操作清单。
        yield from Method.readColumnValues(self.excel_data, self.file_column_data, self.sheet_data)

    def _run_action(self, action_name, action):
        # 所有业务按钮统一处理异常和完成提示。
        try:
            count = action()
        except Exception as exc:
            QMessageBox.critical(self, "错误", f"{action_name}失败：{exc}")
            return

        QMessageBox.information(self, "完成", f"{action_name}完成，共处理 {count} 个文件")

    def refreshSheetOptions(self, show_warning=False):
        if self.sheet_combo is None:
            return

        # 更换表格文件后刷新工作表下拉框，并尽量保留当前选择。
        current_sheet = self.sheet_combo.currentText() or self.sheet_data
        self.sheet_combo.clear()

        excel_path = Path(self.excel_edit.text().strip()) if self.excel_edit else Path(self.excel_data)
        if not excel_path.is_file() or excel_path.suffix.lower() not in EXCEL_EXTENSIONS:
            return

        try:
            sheet_names = self._load_sheet_names(excel_path)
        except Exception as exc:
            if show_warning:
                QMessageBox.warning(self, "警告", f"无法读取工作表：{exc}")
            return

        self.sheet_combo.addItems(sheet_names)
        if not sheet_names:
            return

        if current_sheet in sheet_names:
            self.sheet_combo.setCurrentText(current_sheet)
        else:
            self.sheet_combo.setCurrentIndex(0)

    def chooseExcelFile(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择表格文件",
            self._desktop_path(),
            "Excel 文件 (*.xlsx *.xlsm *.xltx *.xltm);;所有文件 (*.*)",
        )
        if file_path:
            self.excel_edit.setText(file_path)
            self.refreshSheetOptions(show_warning=True)

    def chooseStartFolder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "选择文件起始地址", self._desktop_path())
        if folder_path:
            self.start_path_edit.setText(folder_path)

    def chooseTargetFolder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "选择文件目标地址", self._desktop_path())
        if folder_path:
            self.target_path_edit.setText(folder_path)

    def showConfigPage(self):
        # 回到展示页前重新读取本机保存的配置。
        self.config = Method.readData()
        self._sync_config_fields()
        self._refresh_show_page()
        self.stack.setCurrentWidget(self.show_page)

    def inputDataButton(self):
        # 打开修改页时，把当前配置填入输入控件。
        self.start_path_edit.setText(self.start_path_data)
        self.target_path_edit.setText(self.target_path_data)
        self.excel_edit.setText(self.excel_data)
        self.file_column_edit.setText(self.file_column_data)
        self.original_column_edit.setText(self.original_column_data)
        self.new_column_edit.setText(self.new_column_data)
        self.output_column_edit.setText(self.output_column_data)
        self.refreshSheetOptions()
        if self.sheet_data:
            index = self.sheet_combo.findText(self.sheet_data)
            if index >= 0:
                self.sheet_combo.setCurrentIndex(index)
        self.stack.setCurrentWidget(self.input_page)

    def backButton(self):
        self.stack.setCurrentWidget(self.show_page)

    def executeAction(self):
        # 普通执行：文件改名使用“原名称列 -> 新名称列”。
        action_map = {
            "文件读取": self.readButton,
            "文件复制": self.copyButton,
            "文件移动": self.moveButton,
            "文件替换": self.replaceButton,
            "文件改名": self.renamingButton,
            "文件删除": self.removeButton,
        }
        action = action_map.get(self.action_combo.currentText(), self.readButton)
        action()

    def executeFullNameAction(self):
        # 全名称操作：只有文件改名切换为“文件名列 -> 写入列”，其它操作保持一致。
        action_map = {
            "文件读取": self.readButton,
            "文件复制": self.copyButton,
            "文件移动": self.moveButton,
            "文件替换": self.replaceButton,
            "文件改名": self.renamingFullNameButton,
            "文件删除": self.removeButton,
        }
        action = action_map.get(self.action_combo.currentText(), self.readButton)
        action()

    def initializeSheetButton(self):
        try:
            Method.initializeSheetMethod(
                self.excel_data,
                self.sheet_data,
                self.file_column_data,
                self.original_column_data,
                self.new_column_data,
                self.output_column_data,
            )
        except Exception as exc:
            QMessageBox.critical(self, "错误", f"初始化表格数据失败：{exc}")
            return

        QMessageBox.information(self, "完成", "初始化表格数据完成")

    def readButton(self):
        self._run_action(
            "文件读取",
            lambda: Method.readMethod(
                self.start_path_data,
                self.excel_data,
                self.file_column_data,
                self.original_column_data,
                self.new_column_data,
                self.output_column_data,
                self.sheet_data,
                self.include_subdirs_check.isChecked(),
            ),
        )

    def copyButton(self):
        def action():
            copied_files = []
            for file_name in self._file_names_from_excel():
                copied_files.extend(
                    Method.copyMethod(
                        self.start_path_data,
                        self.target_path_data,
                        file_name,
                        self.include_subdirs_check.isChecked(),
                    )
                )
            return len(copied_files)

        self._run_action("文件复制", action)

    def moveButton(self):
        def action():
            moved_files = []
            for file_name in self._file_names_from_excel():
                moved_files.extend(
                    Method.moveMethod(
                        self.start_path_data,
                        self.target_path_data,
                        file_name,
                        self.include_subdirs_check.isChecked(),
                    )
                )
            return len(moved_files)

        self._run_action("文件移动", action)

    def replaceButton(self):
        def action():
            replaced_files = []
            for file_name in self._file_names_from_excel():
                replaced_files.extend(
                    Method.replaceMethod(
                        self.start_path_data,
                        self.target_path_data,
                        file_name,
                        self.include_subdirs_check.isChecked(),
                    )
                )
            return len(replaced_files)

        self._run_action("文件替换", action)

    def _rename_with_columns(self, start_column, target_column, action_prefix=""):
        # 普通改名和全名称改名共用这里，只替换传入的两列。
        if self.folder_only_check.isChecked():
            action_name = f"{action_prefix}文件夹改名"
            rename_method = Method.renamingFolderMethod
        else:
            action_name = f"{action_prefix}文件改名"
            rename_method = Method.renamingMethod

        self._run_action(
            action_name,
            lambda: len(
                rename_method(
                    self.start_path_data,
                    self.excel_data,
                    start_column,
                    target_column,
                    self.sheet_data,
                    self.include_subdirs_check.isChecked(),
                )
            ),
        )

    def renamingButton(self):
        self._rename_with_columns(self.original_column_data, self.new_column_data)

    def renamingFullNameButton(self):
        self._rename_with_columns(self.file_column_data, self.output_column_data, "全名称")

    def removeButton(self):
        def action():
            removed_files = []
            for file_name in self._file_names_from_excel():
                removed_files.extend(
                    Method.removeMethod(
                        self.start_path_data,
                        file_name,
                        self.include_subdirs_check.isChecked(),
                    )
                )
            return len(removed_files)

        self._run_action("文件删除", action)

    def defineDataButton(self):
        # 确认修改前先做路径、表格和列号校验，再写入本机配置。
        start_path = Path(self.start_path_edit.text().strip())
        target_path = Path(self.target_path_edit.text().strip())
        excel_path = Path(self.excel_edit.text().strip())
        sheet_name = self.sheet_combo.currentText().strip()
        file_column = self.file_column_edit.text().strip()
        original_column = self.original_column_edit.text().strip()
        new_column = self.new_column_edit.text().strip()
        output_column = self.output_column_edit.text().strip()

        if not start_path.is_dir() or not target_path.is_dir():
            QMessageBox.warning(self, "警告", "请输入正确的文件夹地址")
            return

        if not excel_path.is_file() or excel_path.suffix.lower() not in EXCEL_EXTENSIONS:
            QMessageBox.warning(self, "警告", "请使用可读取的 Excel 表格文件")
            return

        try:
            sheet_names = self._load_sheet_names(excel_path)
        except Exception as exc:
            QMessageBox.warning(self, "警告", f"无法打开表格文件：{exc}")
            return

        if sheet_name not in sheet_names:
            QMessageBox.warning(self, "警告", "没有找到工作表")
            return

        columns = {
            "文件名列": file_column,
            "原名称列": original_column,
            "新名称列": new_column,
            "写入列": output_column,
        }
        if any(not column.isdigit() or int(column) < 1 for column in columns.values()):
            QMessageBox.warning(self, "警告", "请输入正确的数据列")
            return

        if len({int(column) for column in columns.values()}) != len(columns):
            QMessageBox.warning(self, "警告", "文件名列、原名称列、新名称列、写入列不能重复")
            return

        self.config.update(
            {
                "start_path": str(start_path),
                "target_path": str(target_path),
                "excel": str(excel_path),
                "sheetname": sheet_name,
                "file_column": file_column,
                "original_column": original_column,
                "new_column": new_column,
                "output_column": output_column,
                "start": file_column,
                "target": new_column,
            }
        )
        Method.storeData(self.config)
        self.showConfigPage()
