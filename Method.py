import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from PySide6.QtCore import QSettings

from config_defaults import DEFAULT_CONFIG


SETTINGS_ORG = "WenjianChuli"
SETTINGS_APP = "FileProcessor"
# 只处理 A1、$A$1 这类单元格引用，供简单拼接公式解析使用。
CELL_REFERENCE_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?(\d+)$")


def _settings():
    # QSettings 会把用户选择的路径和列号保存到当前系统用户配置里。
    return QSettings(SETTINGS_ORG, SETTINGS_APP)


class Method:
    @staticmethod
    def _safe_child_path(start_path, file_name):
        # 防止表格里的文件名写成绝对路径或跳出起始目录。
        start_dir = Path(start_path).resolve()
        file_text = str(file_name).strip()
        if not file_text:
            return None

        relative_file = Path(file_text)
        if relative_file.is_absolute():
            return None

        child_path = (start_dir / relative_file).resolve()
        if not child_path.is_relative_to(start_dir):
            return None

        return child_path

    @staticmethod
    def getSelectedFiles(start_path, include_subdirs=False):
        start_dir = Path(start_path)
        if include_subdirs:
            return [path for path in start_dir.rglob("*") if path.is_file()]

        return [path for path in start_dir.iterdir() if path.is_file()]

    @staticmethod
    def getSelectedFolders(start_path, include_subdirs=False):
        start_dir = Path(start_path)
        if include_subdirs:
            return [path for path in start_dir.rglob("*") if path.is_dir()]

        return [path for path in start_dir.iterdir() if path.is_dir()]

    @staticmethod
    def _file_name_only(file_name):
        return Path(str(file_name).strip()).name

    @staticmethod
    def _split_concat_formula(expression):
        # 拆分 Excel 的 & 拼接公式，同时避开引号里的 & 字符。
        parts = []
        current = []
        in_quote = False
        index = 0

        while index < len(expression):
            char = expression[index]
            current.append(char)

            if char == '"':
                if in_quote and index + 1 < len(expression) and expression[index + 1] == '"':
                    index += 1
                    current.append(expression[index])
                else:
                    in_quote = not in_quote
            elif char == "&" and not in_quote:
                current.pop()
                parts.append("".join(current))
                current = []

            index += 1

        parts.append("".join(current))
        return parts

    @staticmethod
    def _cell_value(sheet, values_sheet, row, column, seen=None):
        # 优先自己计算本工具生成的拼接公式，算不了时再读取 Excel 缓存值。
        value = sheet.cell(row, column).value
        if isinstance(value, str) and value.startswith("="):
            evaluated_value = Method._evaluate_concat_formula(value, sheet, values_sheet, seen or set())
            if evaluated_value is not None:
                return evaluated_value

            return values_sheet.cell(row, column).value

        cached_value = values_sheet.cell(row, column).value
        if cached_value is not None:
            return cached_value

        if not isinstance(value, str) or not value.startswith("="):
            return value

        return None

    @staticmethod
    def _evaluate_concat_formula(formula, sheet, values_sheet, seen):
        # 这里只支持“单元格 & 文本 & 单元格”这类简单公式，足够覆盖文件名拼接。
        parts = Method._split_concat_formula(formula[1:].strip())
        values = []

        for part in parts:
            token = part.strip()
            if len(token) >= 2 and token[0] == '"' and token[-1] == '"':
                values.append(token[1:-1].replace('""', '"'))
                continue

            match = CELL_REFERENCE_RE.match(token)
            if not match:
                return None

            column = column_index_from_string(match.group(1))
            row = int(match.group(2))
            cell_key = (sheet.title, row, column)
            # 避免公式互相引用时无限递归。
            if cell_key in seen:
                return None

            seen.add(cell_key)
            value = Method._cell_value(sheet, values_sheet, row, column, seen)
            seen.remove(cell_key)
            values.append("" if value is None else str(value))

        return "".join(values)

    @staticmethod
    def readColumnValues(excel_path, column, sheetname):
        # 复制、移动、删除会复用这里读取“文件名列”的有效文件名。
        workbook = load_workbook(excel_path, data_only=False)
        values_workbook = load_workbook(excel_path, data_only=True)
        try:
            sheet = workbook[sheetname]
            values_sheet = values_workbook[sheetname]
            target_column = int(column)
            values = []

            for line in range(2, sheet.max_row + 1):
                value = Method._cell_value(sheet, values_sheet, line, target_column)
                if value is None:
                    continue

                text = str(value).strip()
                if text:
                    values.append(text)

            return values
        finally:
            workbook.close()
            values_workbook.close()

    @staticmethod
    def _read_name_pairs(excel_path, start, target, sheetname):
        # 改名操作按两列读取映射关系：旧内容 -> 新内容。
        workbook = load_workbook(excel_path, data_only=False)
        values_workbook = load_workbook(excel_path, data_only=True)
        try:
            sheet = workbook[sheetname]
            values_sheet = values_workbook[sheetname]
            start_column = int(start)
            target_column = int(target)
            name_pairs = []

            for line in range(2, sheet.max_row + 1):
                start_value = Method._cell_value(sheet, values_sheet, line, start_column)
                target_value = Method._cell_value(sheet, values_sheet, line, target_column)
                if start_value is None or target_value is None:
                    continue

                start_name = str(start_value).strip()
                target_name = str(target_value).strip()
                if start_name and target_name:
                    name_pairs.append((start_name, target_name))

            return name_pairs
        finally:
            workbook.close()
            values_workbook.close()

    @staticmethod
    def _rename_selected_paths(selected_paths, start_path, name_pairs, include_subdirs=False):
        # 支持两种匹配：文件名内局部替换，或包含子目录时按相对路径完整匹配。
        renamed_paths = []
        start_dir = Path(start_path)

        for path in selected_paths:
            relative_name = str(path.relative_to(start_dir))
            relative_name_alt = relative_name.replace("\\", "/")

            for start_name, target_name in name_pairs:
                new_name = None

                if start_name in path.name:
                    new_name = path.name.replace(start_name, Method._file_name_only(target_name))
                elif include_subdirs and start_name in {relative_name, relative_name_alt}:
                    new_name = Method._file_name_only(target_name)

                if not new_name:
                    continue

                new_path = path.with_name(new_name)
                if new_path.exists():
                    continue

                path.rename(new_path)
                renamed_paths.append(str(new_path))
                break

        return renamed_paths

    @staticmethod
    def _find_files(start_path, file_name, include_subdirs=False):
        # 先按表格里的相对路径精确查找，再按文件名在子目录中搜索。
        start_dir = Path(start_path).resolve()
        file_text = str(file_name).strip()
        if not file_text:
            return []

        direct_file = Method._safe_child_path(start_dir, file_text)
        if direct_file and direct_file.is_file():
            return [direct_file]

        if include_subdirs:
            return [path for path in Method.getSelectedFiles(start_dir, True) if path.name == file_text]

        return []

    @staticmethod
    def _copy_or_move(start_path, over_path, file_name, include_subdirs=False, move=False):
        # 复制和移动共享同一套定位逻辑，包含子目录时保留相对目录结构。
        start_dir = Path(start_path).resolve()
        target_dir = Path(over_path)
        handled_files = []

        for source_file in Method._find_files(start_path, file_name, include_subdirs):
            if include_subdirs:
                target_name = source_file.relative_to(start_dir)
            else:
                target_name = source_file.name

            target_file = target_dir / target_name
            if target_file.exists():
                continue

            target_file.parent.mkdir(parents=True, exist_ok=True)
            if move:
                shutil.move(str(source_file), str(target_file))
            else:
                shutil.copy2(source_file, target_file)

            handled_files.append(str(target_file))

        return handled_files

    @staticmethod
    def _target_files_by_name(target_path, file_name):
        target_dir = Path(target_path)
        return [path for path in target_dir.rglob("*") if path.is_file() and path.name == file_name]

    # 文件写入方法
    @staticmethod
    def readMethod(
        start_path,
        excel_path,
        file_column,
        original_column,
        new_column,
        output_column,
        sheetname,
        include_subdirs=False,
    ):
        workbook = load_workbook(excel_path)
        try:
            sheet = workbook[sheetname]
            file_column = int(file_column)
            original_column = int(original_column)
            new_column = int(new_column)
            output_column = int(output_column)
            start_dir = Path(start_path)
            count = 0

            # 只清空本工具负责的四列，保留表格里其它辅助列。
            for column in {file_column, original_column, new_column, output_column}:
                for line in range(2, sheet.max_row + 1):
                    sheet.cell(line, column).value = None

            for row_index, file_path in enumerate(Method.getSelectedFiles(start_dir, include_subdirs), start=2):
                original_name = file_path.stem
                original_name_cell = f"{get_column_letter(original_column)}{row_index}"
                new_name_cell = f"{get_column_letter(new_column)}{row_index}"

                # 文件名列和写入列都写公式，便于用户只改名称部分时自动保留原后缀。
                sheet.cell(row_index, file_column).value = f'={original_name_cell}&"{file_path.suffix}"'
                sheet.cell(row_index, original_column).value = original_name
                sheet.cell(row_index, new_column).value = original_name
                sheet.cell(row_index, output_column).value = f'={new_name_cell}&"{file_path.suffix}"'
                count += 1

            workbook.save(excel_path)
            return count
        finally:
            workbook.close()

    # 初始化表格数据
    @staticmethod
    def initializeSheetMethod(excel_path, sheetname, file_column, original_column, new_column, output_column):
        workbook = load_workbook(excel_path)
        try:
            sheet = workbook[sheetname]
            file_column = int(file_column)
            original_column = int(original_column)
            new_column = int(new_column)
            output_column = int(output_column)

            # 初始化会重置整张工作表，用于重新生成标准标题和空表。
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None

            sheet.cell(1, file_column).value = "文件名列"
            sheet.cell(1, original_column).value = "原名称列"
            sheet.cell(1, new_column).value = "新名称列"
            sheet.cell(1, output_column).value = "写入列"

            workbook.save(excel_path)
        finally:
            workbook.close()

    # 文件复制方法
    @staticmethod
    def copyMethod(start_path, over_path, file_name, include_subdirs=False):
        return Method._copy_or_move(start_path, over_path, file_name, include_subdirs, move=False)

    # 文件移动方法
    @staticmethod
    def moveMethod(start_path, over_path, file_name, include_subdirs=False):
        return Method._copy_or_move(start_path, over_path, file_name, include_subdirs, move=True)

    # 文件替换方法
    @staticmethod
    def replaceMethod(start_path, over_path, file_name, include_subdirs=False):
        source_files = Method._find_files(start_path, file_name, include_subdirs)
        if len(source_files) > 1:
            raise ValueError(f"起始地址中找到多个同名源文件，无法判断使用哪一个替换：{file_name}")

        if not source_files:
            return []

        source_file = source_files[0]
        source_path = source_file.resolve()
        replaced_files = []

        for target_file in Method._target_files_by_name(over_path, source_file.name):
            if target_file.resolve() == source_path:
                continue

            shutil.copy2(source_file, target_file)
            replaced_files.append(str(target_file))

        return replaced_files

    # 文件改名方法
    @staticmethod
    def renamingMethod(start_path, excel_path, start, target, sheetname, include_subdirs=False):
        name_pairs = Method._read_name_pairs(excel_path, start, target, sheetname)
        selected_files = Method.getSelectedFiles(start_path, include_subdirs)
        return Method._rename_selected_paths(selected_files, start_path, name_pairs, include_subdirs)

    # 文件夹改名方法
    @staticmethod
    def renamingFolderMethod(start_path, excel_path, start, target, sheetname, include_subdirs=False):
        name_pairs = Method._read_name_pairs(excel_path, start, target, sheetname)
        selected_folders = Method.getSelectedFolders(start_path, include_subdirs)
        selected_folders.sort(key=lambda path: len(path.parts), reverse=True)
        return Method._rename_selected_paths(selected_folders, start_path, name_pairs, include_subdirs)

    # 文件删除方法
    @staticmethod
    def removeMethod(start_path, file_name, include_subdirs=False):
        removed_files = []

        for file_path in Method._find_files(start_path, file_name, include_subdirs):
            file_path.unlink()
            removed_files.append(str(file_path))

        return removed_files

    # 读取本机缓存数据
    @staticmethod
    def readData():
        settings = _settings()
        config = DEFAULT_CONFIG.copy()

        for key, default_value in DEFAULT_CONFIG.items():
            value = settings.value(key, default_value)
            config[key] = default_value if value is None else str(value)

        config["file_column"] = str(config.get("file_column") or config.get("start") or "1")
        config["original_column"] = str(config.get("original_column") or config.get("target") or "2")
        config["new_column"] = str(config.get("new_column") or "3")
        config["output_column"] = str(config.get("output_column") or "4")
        # 兼容旧配置字段：start/target 仍保持为文件名列/新名称列。
        config["start"] = config["file_column"]
        config["target"] = config["new_column"]
        return config

    # 写入本机缓存数据
    @staticmethod
    def storeData(data):
        settings = _settings()
        config = {**DEFAULT_CONFIG, **data}

        for key in DEFAULT_CONFIG:
            value = config.get(key, "")
            settings.setValue(key, "" if value is None else str(value))

        settings.sync()
